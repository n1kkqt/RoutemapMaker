import sys
import io 
import os
from miette import DocReader
from PyQt5.QtWidgets import (QWidget, QLabel, QLineEdit, QPushButton, QApplication)
import docx
from openpyxl import load_workbook
import shutil
from os import listdir
from os.path import isfile, join
from openpyxl.styles import Border, Side


def replace_trash(strin):
    with io.open('strfile.txt', mode="w", encoding="utf_8_sig", ) as fd: 
        fd.write(strin)
    with io.open('strfile.txt', mode="r", encoding="utf_8_sig") as fd: 
        strin = fd.read() 
    os.remove('strfile.txt')
    return strin

def getText(filename):
    doc = docx.Document(filename)
    fullText = []
    for para in doc.paragraphs:
        fullText.append(para.text)
    return '\n'.join(fullText)

def f1(slist):
    slist = [s and s.strip() for s in slist]
    return list(filter(None, slist))

def getData(filename):
    if filename[-4:-1]=='doc':
        partext=getText(filename).split("\n")
    elif filename[-4:-1]=='.do':
        doc = DocReader(filename)
        partext=replace_trash(doc.read().decode("utf-8", "strict").replace("","").replace("","")).split("\n")
    data=[]
    for i in range(0,len(partext)):
        if partext[i].find('В  ')!=-1:
            data.append(partext[i])
    for i in range(0,len(data)):
        data[i]=data[i].replace('В',"")
        data[i]=data[i].split('  ')
        data[i]=list(filter(None,data[i]))
    return data       

class Window(QWidget):

    def __init__(self):
        super().__init__()
        self.initUI()


    def initUI(self):

        self.lbl = QLabel(self)
        self.lbl.move(0, 0)
        self.lbl.setText('Путь к папке с 2 пустыми образцами .xlsx(с одним и двумя листами)')
        
        self.xlsxt = QLineEdit(self)
        self.xlsxt.resize(350,20)
        self.xlsxt.move(0, 15)
        
        self.lbl2 = QLabel(self)
        self.lbl2.move(0, 35)
        self.lbl2.setText('Путь к папке с техпроцессами')
        
        self.openpath = QLineEdit(self)
        self.openpath.resize(350,20)
        self.openpath.move(0, 50)
        
        self.lbl3 = QLabel(self)
        self.lbl3.move(0, 70)
        self.lbl3.setText('Куда сохранять')
        
        self.savepath = QLineEdit(self)
        self.savepath.resize(350,20)
        self.savepath.move(0, 85)
        
        self.lbl4 = QLabel(self)
        self.lbl4.move(0, 105)
        self.lbl4.setText('Разработал')
        
        self.developer = QLineEdit(self)
        self.developer.resize(150,20)
        self.developer.move(0, 120)
        
        self.lbl5 = QLabel(self)
        self.lbl5.move(0, 140)
        self.lbl5.setText('Проверил')
        
        self.check = QLineEdit(self)
        self.check.resize(150,20)
        self.check.move(0, 155)
        
        self.lbl7 = QLabel(self)
        self.lbl7.move(180, 140)
        self.lbl7.setText('Название техпроцесса')
        
        self.tech = QLineEdit(self)
        self.tech.resize(150,20)
        self.tech.move(180, 155)
        
        self.lbl6 = QLabel(self)
        self.lbl6.move(0, 175)
        self.lbl6.setText('Утвердил')
        
        self.approve = QLineEdit(self)
        self.approve.resize(150,20)
        self.approve.move(0, 190)
        
        self.btn = QPushButton('Создать', self)
        self.btn.move(0, 230)
        self.btn.clicked.connect(self.go)
        
        self.btnex = QPushButton('Выход', self)
        self.btnex.move(0, 250)
        self.btnex.clicked.connect(self.closeapp)

        self.setGeometry(400, 400, 355, 270)
        self.setWindowTitle('RoutemapMaker by Nikita Lokhmachev')
        self.show()

    def go(self):
        xlsxtval=self.xlsxt.text()
        openpathval=self.openpath.text()
        savedfiles=self.savepath.text()+'\\'
        devval=self.developer.text()
        checkval=self.check.text()
        appval=self.approve.text()
        techproc=self.tech.text()
        onlyfiles = [f for f in listdir(openpathval) if isfile(join(openpathval, f))]
        onlyfilesxlsx = [f for f in listdir(xlsxtval) if isfile(join(xlsxtval, f))]
        xlsxtval1=xlsxtval+'\\'+onlyfilesxlsx[0]
        xlsxtval2=xlsxtval+'\\'+onlyfilesxlsx[1]
        for i in range(0,len(onlyfiles)):
            datatemp=getData(openpathval+'\\'+onlyfiles[i])
            if onlyfiles[i][-4:-1]=='doc':
                onlyfiles[i]=onlyfiles[i].replace('.docx','.xlsx')
            elif onlyfiles[i][-4:-1]=='.do':
                onlyfiles[i]=onlyfiles[i].replace('.doc','.xlsx')
                datatemp[-4:-1]=[]
                del datatemp[-1]
                for x in range(0,len(datatemp)):
                    datatemp[x]=f1(datatemp[x])
            newfullname=savedfiles+onlyfiles[i]
            shutil.copyfile(xlsxtval1, newfullname )
            if len(datatemp)>19:
                routemap=load_workbook(xlsxtval2)
            else: 
                routemap=load_workbook(xlsxtval2)
            ws=routemap['Лист1']
            for j in range(0,len(datatemp)):
                if j<19:
                    tempstr1='A'+str(j+10)
                    tempstr2='B'+str(j+10)
                    tempstr3='D'+str(j+10)
                    ws[tempstr1]=datatemp[j][1]
                    ws[tempstr2]=datatemp[j][0]
                    ws[tempstr3]=datatemp[j][2]
                else:
                    tempstr1='A'+str(j+23)
                    tempstr2='B'+str(j+23)
                    tempstr3='D'+str(j+23)
                    ws[tempstr1]=datatemp[j][1]
                    ws[tempstr2]=datatemp[j][0]
                    ws[tempstr3]=datatemp[j][2]
                kgdp=onlyfiles[i].replace('.xlsx','')
                ws['D3']=kgdp
                if len(datatemp)>19:
                    ws['D35']=kgdp
            ws['K30']=devval
            ws['K31']=checkval
            ws['K32']=appval
            ws['K3']=techproc
            for x in range(1,15):
                for y in range(1,9):
                    ws.cell(y,x).border = Border(top = Side(border_style='medium', color='FF000000'),    
                                          right = Side(border_style='medium', color='FF000000'), 
                                          bottom = Side(border_style='medium', color='FF000000'),
                                          left = Side(border_style='medium', color='FF000000'))
            for x in range(1,15):
                for y in range(10, 29):
                    ws.cell(y,x).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                          right = Side(border_style='thin', color='FF000000'), 
                                          bottom = Side(border_style='thin', color='FF000000'),
                                          left = Side(border_style='thin', color='FF000000'))
            for x in range(9,15):
                for y in range(29,33):
                    ws.cell(y,x).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                          right = Side(border_style='thin', color='FF000000'), 
                                          bottom = Side(border_style='thin', color='FF000000'),
                                          left = Side(border_style='thin', color='FF000000'))
            if len(datatemp)>19:
                ws['K35']=techproc
                for x in range(1,15):
                    for y in range(35,41):
                        ws.cell(y,x).border = Border(top = Side(border_style='medium', color='FF000000'),    
                                              right = Side(border_style='medium', color='FF000000'), 
                                              bottom = Side(border_style='medium', color='FF000000'),
                                              left = Side(border_style='medium', color='FF000000'))
                for x in range(1,15):
                    for y in range(42, 71):
                        ws.cell(y,x).border = Border(top = Side(border_style='thin', color='FF000000'),    
                                              right = Side(border_style='thin', color='FF000000'), 
                                              bottom = Side(border_style='thin', color='FF000000'),
                                              left = Side(border_style='thin', color='FF000000'))
            routemap.save(newfullname)

    def closeapp(self):
        sys.exit()

if __name__ == '__main__':

    app = QApplication(sys.argv)
    ex = Window()
    sys.exit(app.exec_())
